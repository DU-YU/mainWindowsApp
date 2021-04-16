# -*- coding: utf-8 -*-

import sys
import threading
import time
# from ctypes import windll, byref
# from ctypes.wintypes import POINT
# from turtle import onrelease

# from PyQt5.QtGui import QKeyEvent
# from keyboard import on_press, on_release
from PyQt5 import QtWidgets
from pynput import keyboard
import operator
# from ctypes import sizeof
from PyQt5.QtCore import QThread, pyqtSignal, QMutex

import openpyxl
import re
# import binascii

# import win32api
import win32con
import win32gui

# from PyQt5 import QtGui, QtCore
from openpyxl.utils import get_column_letter
# from xlrd import open_workbook, empty_cell
# import cv2
from PyQt5.QtCore import pyqtSlot
from PyQt5.QtWidgets import QApplication, QMainWindow, QFileDialog
# from win32file import WriteFile
import datetime

from win32event import WaitForSingleObject

from ui_MainWindow import Ui_MainWindow

# from myMainWindow import QmyMainWindow

# from PyQt5.QtWidgets import
# from PyQt5.QtGui import
# from PyQt5.QtSql import
# from PyQt5.QtMultimedia import
# from PyQt5.QtMultimediaWidgets import

# ==============全局变量========================

awb_list = []
count = 0
time_interval = 300
awb_handle = 0
cnt = 0

# The currently active modifiers
current = set()

lock = threading.Lock()

COMBINATIONS = [
    # { keyboard.Key.shift, keyboard.Key.ctrl_l },
    # { keyboard.Key.shift_r, keyboard.Key.ctrl_r },
    {keyboard.Key.alt_l, keyboard.KeyCode(char='l')},
    {keyboard.Key.alt_gr, keyboard.KeyCode(char='l')},
    {keyboard.Key.alt_l, keyboard.KeyCode(char='r')},
    {keyboard.Key.alt_gr, keyboard.KeyCode(char='r')},
    # { keyboard.Key.esc }
]


def regExCheck(awb):
    # print(awb)
    try:
        result = re.match(r'(888|666|555[\d]{9})|(J[A-Z]{1}[\d]{10})|(1[\d]{9})|(Y|F|LY[\d]{8})', awb, flags=0)
    except[]:
        pass
    # print(result)
    if result is None:
        # print(False)
        return False
    else:
        # print(True)
        return True


def repeat_thread_detection(tName):
    # 判断 tName线程是否处于活动状态
    for item in threading.enumerate():
        if tName == item.name:  # 如果名字相同，说明tName线程在活动的线程列表里面
            # print(tName)
            return True
    return False


class QmyMainWindow(QMainWindow):

    def __init__(self, parent=None):
        super().__init__(parent)  # 调用父类构造函数，创建窗体
        self.ui = Ui_MainWindow()  # 创建UI对象
        self.ui.setupUi(self)  # 构造UI界面
        # 实例化线程对象
        self.work = WorkThread()
        self.handle = getWindowsHandleThread()
        # self.execute()
        # self.run()
        self.ui.startPushButton.toggled.connect(self.execute)
        self.ui.getHandlePushButton.clicked.connect(self.run)
        QtWidgets.QMessageBox.about(self, "Tips", "This App is used to check lost scanning, which can't replace pickup."
                                                  " by manually. And this app will be expired at 31st December 2021. "
                                                  "\t\t\t\t\t\t\t\t\t\t\t\t\t\t\t\t\t———— SUPPORTED BY JKT-STAR")
        now_time = datetime.datetime.now().strftime('%Y-%m-%d')
        # delta = int(now_time)-int('2021-04-14')
        # print(delta)
        if now_time > '2021-04-16':
            sys.exit()
        else:
            pass
    # ==============自定义功能函数========================
    letter = 'A'
    cnt = 0
    # 自定义信号对象。参数str就代表这个信号可以传一个字符串
    # trigger = pyqtSignal(str)

    def execute(self):
        if self.ui.startPushButton.isChecked():
            # if not self.repeat_thread_detection('Dummy-2'):
            # threading.Thread(target=self.display, name='Dummy-2').start()
            # 启动线程
            self.work.start()
            # 线程自定义信号连接的槽函数
            try:
                #只让执行一次
                if self.cnt == 0:
                    self.work.trigger.connect(self.display)
                    self.cnt += 1
            except[]:
                pass
        else:
            pass
            # self.work.quit()

    def display(self, str):
        print(threading.enumerate())
        if self.ui.startPushButton.isChecked() is True:
            # 由于自定义信号时自动传递一个字符串参数，所以在这个槽函数中要接受一个参数
            self.ui.awbLineEdit.setText(str)
            self.ui.countsLcdNumber.display(count)
            win32gui.SendMessage(awb_handle, win32con.WM_SETTEXT, 'Edit', str)
            win32gui.PostMessage(awb_handle, win32con.WM_KEYDOWN, win32con.VK_RETURN, 0)
            cursor = self.ui.textEdit2.textCursor()
            # cursor.movePosition(QtGui.QTextCursor.End)
            cursor.insertText(str)
            cursor.insertText('\n')
            self.ui.textEdit2.setTextCursor(cursor)
            self.ui.textEdit2.ensureCursorVisible()
            # self.work.wait()
            print(str)
        else:
            self.work.wait()
            pass

    def run(self):
        self.handle.start()
        self.handle.trigger.connect(self.displayHandle)

    def displayHandle(self, flag):
        if flag:
            position = win32gui.GetCursorPos()
            awbhandle = win32gui.WindowFromPoint(position)
            self.ui.awbHandleLineEdit.setText(str(awbhandle))
            handlename = win32gui.GetWindowText(awbhandle)
            self.ui.awbHandleNameLineEdit.setText(handlename)
            self.ui.awbHandleLineEdit.setEnabled(False)
            self.handle.quit()
        else:
            self.ui.awbHandleLineEdit.setEnabled(True)
            self.ui.awbHandleNameLineEdit.setText("")
            self.handle.quit()

    # @pyqtSlot()
    # def focusOutput(self, hwnd, ms):
    # print("test")
    # #for awb in awb_list:
    # self.ui.awbLineEdit.setText(awb)
    # win32gui.SendMessage(hwnd, win32con.WM_SETTEXT, 'Edit', awb)
    # win32gui.SendMessage(hwnd, win32con.WM_CHAR, awb, 0)
    # win32gui.SendMessage(hwnd, win32con.WM_CHAR, '\r\n', 0)
    # win32gui.PostMessage(hwnd, win32con.WM_KEYDOWN, win32con.VK_RETURN, 0)
    # time.sleep(ms)

    @pyqtSlot()
    def on_openFilePushButton_clicked(self):
        directory, attribute = QFileDialog.getOpenFileName(self, "open file dialog", "C:\\Users\\XU\\DeskTop\\",
                                                           "Excel(*xlsx);;CvsFiles(*.cvs)")
        self.ui.lineEdit.setText(directory)
        # print(directory)
        global awb_list
        awb_list = []
        self.ui.textEdit.clear()
        if attribute == "Excel(*xlsx)":
            wb = openpyxl.load_workbook(directory)
            sheets = wb.get_sheet_names()
            ws = wb.get_sheet_by_name(sheets[0])
            for row in ws.iter_rows(min_row=1, max_row=3, max_col=6):  # 打印1-3行，1-6列中的内容
                for cell in row:
                    if cell.value is not None:
                        # print(cell.value)
                        # print(cell.column)
                        # letter = get_column_letter(cell.column)
                        # print(letter)
                        if regExCheck(str(cell.value)) is True:
                            letter = get_column_letter(cell.column)
                            # awb_column = cell.column
                            # print(letter)
                            # print("OK")
                            break
                        else:
                            pass
            col_range = ws[letter:letter]
            # print(col_range)
            for cell in col_range[1:]:  # 打印BC两列单元格中的值内容
                awb_list.append(str(cell.value))
                cursor = self.ui.textEdit.textCursor()
                # cursor.movePosition(QtGui.QTextCursor.End)
                cursor.insertText(str(cell.value))
                cursor.insertText('\n')
                self.ui.textEdit.setTextCursor(cursor)
                self.ui.textEdit.ensureCursorVisible()
        # print(len(awb_list))

    @pyqtSlot()
    def on_getHandlePushButton_clicked(self):
        QtWidgets.QMessageBox.about(self, "TIPS",
                                    "Please put your mouse cursor on awb input box, then press \"Alt+L\" to lock awb"
                                    "handle. And handle is not NULL. \"Alt+R\" can unlock awb handle.")

    def on_startPushButton_toggled(self, checked):
        if len(self.ui.awbHandleLineEdit.text()) == 0:
            QtWidgets.QMessageBox.about(self, "WARNING", "AWB Handle is not assigned, Please input or use \"AlT+L\" to "
                                                         "get AWB Handle.")
            return
        # print(threading.enumerate())
        # print(thread_num)
        if self.ui.startPushButton.isChecked():
            # cursor = self.ui.textEdit2.textCursor()
            # cursor.insertText('')
            # self.ui.textEdit2.setTextCursor(cursor)
            # self.ui.textEdit2.ensureCursorVisible()
            self.ui.textEdit2.clear()
            # print(self.ui.textEdit2.getContentsMargins())
            global count
            count = 0
            self.ui.countsLcdNumber.display(count)
            global time_interval
            time_interval = int(self.ui.TimeIntervalLineEdit.text())
            global awb_handle
            awb_handle = int(self.ui.awbHandleLineEdit.text())
            self.ui.awbHandleLineEdit.setEnabled(False)
            self.ui.TimeIntervalLineEdit.setEnabled(False)
            # delay_ms = float(self.ui.TimeIntervalLineEdit.text()) / 1000
            # print(awb_handle)
            # self.focusOutput(awb_handle, delay_ms)
            # win32gui.SendMessage(0x00D05C2, win32con.WM_SETTEXT, 'TRzEdit', 'JP1234567890')
            # win32gui.PostMessage(0x00D05C2, win32con.WM_KEYDOWN, win32con.VK_RETURN, 0)
            self.ui.startPushButton.setText("STOP")
        else:
            self.ui.startPushButton.setText("START")
            self.ui.awbHandleLineEdit.setEnabled(True)
            self.ui.TimeIntervalLineEdit.setEnabled(True)
            # print(False)


#  ==============event处理函数==========================
class WorkThread(QThread):
    # 自定义信号对象。参数str就代表这个信号可以传一个字符串
    trigger = pyqtSignal(str)

    def __int__(self):
        # 初始化函数
        super(WorkThread, self).__init__()

    def run(self):
        # 重写线程执行的run函数
        # 触发自定义信号
        # qmut_1.lock()  # 加锁
        for awb in awb_list:
            if regExCheck(awb):
                if time_interval == 0:
                    time.sleep(0.3)
                else:
                    time.sleep(float(time_interval / 1000))
                global count
                count += 1
                # 通过自定义信号把待显示的字符串传递给槽函数
                self.trigger.emit(awb)
                # print(awb)
            else:
                pass
        # qmut_1.unlock()  # 加锁


class getWindowsHandleThread(QThread):
    # 自定义信号对象。参数str就代表这个信号可以传一个字符串
    trigger = pyqtSignal(int)

    def __int__(self):
        # 初始化函数
        super(getWindowsHandleThread, self).__init__()

    def run(self):
        with keyboard.Listener(on_press=self.on_press, on_release=self.on_release) as listener:
            listener.join()

    def on_press(self, key):
        # print(key)
        if any([key in COMBO for COMBO in COMBINATIONS]):
            global current
            current.add(key)
            if any(all(k in current for k in COMBO) for COMBO in COMBINATIONS):
                # print(current)
                # print(COMBINATIONS[0])
                if operator.eq(current, COMBINATIONS[0]) or operator.eq(current, COMBINATIONS[1]):
                    # print(current)
                    self.trigger.emit(True)
                elif operator.eq(current, COMBINATIONS[2]) or operator.eq(current, COMBINATIONS[3]):
                    # print(current)
                    self.trigger.emit(False)
                else:
                    pass
                # print(operator.eq(current, COMBINATIONS[1]))

    def on_release(self, key):
        try:
            current.remove(key)
        except KeyError:
            pass


#  ==========由connectSlotsByName()自动连接的槽函数============


#  =============自定义槽函数===============================


#  ============窗体测试程序 ================================
if __name__ == "__main__":  # 用于当前窗体测试
    app = QApplication(sys.argv)  # 创建GUI应用程序
    form = QmyMainWindow()  # 创建窗体
    form.show()
    # with keyboard.Listener(on_press=on_press, on_release=on_release) as listener:
    #    listener.join()
    sys.exit(app.exec_())
